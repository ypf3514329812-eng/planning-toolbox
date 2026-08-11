"""Tests for student-friendly Excel, PDF and PNG result exports."""

from pathlib import Path

from matplotlib.figure import Figure
from openpyxl import load_workbook

from planning_toolbox.gui.result_export import export_result_artifacts


class _PreviewStub:
    def __init__(self):
        self.figure = Figure(figsize=(3, 2))
        axis = self.figure.add_subplot(111)
        axis.plot([0, 1], [0, 1])
        axis.set_title("Preview")


def test_export_result_artifacts_creates_excel_pdf_and_png(tmp_path):
    source = tmp_path / "source.dxf"
    source.write_text("source", encoding="utf-8")
    result = {
        "task_type": "indicator",
        "source_file": str(source),
        "source_sha256": "abc123",
        "floors": 6,
        "indicators": [{"parcel_id": "P001", "far": 1.5, "green_ratio_pct": 30}],
        "output_files": [("指标报告", str(tmp_path / "report.txt"))],
    }

    exported = export_result_artifacts(tmp_path, "规划指标计算", result, _PreviewStub())
    paths = {label: path for label, path in exported}

    assert paths["Excel 结果表"].suffix == ".xlsx"
    assert paths["PDF 结果报告"].suffix == ".pdf"
    assert paths["PNG 预览图"].suffix == ".png"
    assert all(path.is_file() and path.stat().st_size > 0 for path in paths.values())

    workbook = load_workbook(paths["Excel 结果表"], read_only=True)
    assert "结果摘要" in workbook.sheetnames
    assert "指标明细" in workbook.sheetnames
    assert workbook["结果摘要"]["B2"].value == "规划指标计算"
