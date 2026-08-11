"""Tests for multi-scenario comparison from saved project files."""

from pathlib import Path

import ezdxf
from matplotlib.figure import Figure
from openpyxl import load_workbook

from planning_toolbox.gui.comparison import (
    build_comparison_row,
    build_comparison_rows,
    export_comparison_csv,
    export_comparison_excel,
)
from planning_toolbox.gui.project_file import save_project
from planning_toolbox.gui.overlay import export_overlay_png, render_project_overlays


def _write_project(path: Path, name: str, result: dict) -> Path:
    return save_project(
        path,
        {
            "dxf_path": f"C:/study/{name}.dxf",
            "output_dir": "C:/study/output",
            "task": {"current_tab_index": 5},
            "last_task_name": name,
            "last_result": result,
        },
    )


def test_comparison_extracts_indicator_and_concept_metrics(tmp_path):
    indicator = _write_project(
        tmp_path / "scheme_a.ptx",
        "规划指标方案 A",
        {
            "task_type": "indicator",
            "indicators": [
                {"site_area_m2": 1000, "far": 1.2, "building_density_pct": 25, "green_ratio_pct": 30},
                {"site_area_m2": 2000, "far": 1.6, "building_density_pct": 35, "green_ratio_pct": 40},
            ],
            "output_files": [],
        },
    )
    concept = _write_project(
        tmp_path / "scheme_b.ptx",
        "概念方案 B",
        {
            "task_type": "concept_plan",
            "parcels_count": 2,
            "parcel_area_m2": 5000,
            "estimated_far": 1.8,
            "actual_coverage_ratio": 0.28,
            "green_area_m2": 1500,
            "parking_generated": 8,
            "parking_required": 10,
            "minimum_setback_m": 5.5,
            "output_files": [],
        },
    )

    row_a = build_comparison_row(indicator)
    rows, errors = build_comparison_rows([indicator, concept])

    assert not errors
    assert len(rows) == 2
    assert row_a["总用地面积(m²)"] == 3000
    assert row_a["FAR"] == "1.20 ~ 1.60"
    assert row_a["建筑密度(%)"] == 30
    assert rows[1]["停车位"] == "8/10"
    assert rows[1]["退线/标准化结果"] == "最小退线 5.50m"


def test_comparison_exports_csv_and_excel(tmp_path):
    rows = [{
        "方案名称": "方案 A",
        "任务类型": "规划指标计算",
        "输入图纸": "a.dxf",
        "有效地块/处理数": 2,
        "总用地面积(m²)": 3000,
        "FAR": "1.20 ~ 1.60",
        "建筑密度(%)": 30,
        "绿地率(%)": 35,
        "停车位": "—",
        "退线/标准化结果": "—",
        "项目文件": "a.ptx",
    }]

    csv_path = export_comparison_csv(tmp_path, rows)
    excel_path = export_comparison_excel(tmp_path, rows)

    assert csv_path.is_file()
    assert excel_path.is_file()
    assert "方案 A" in csv_path.read_text(encoding="utf-8-sig")
    workbook = load_workbook(excel_path, read_only=True)
    assert workbook.sheetnames == ["方案对比"]
    assert workbook["方案对比"]["A2"].value == "方案 A"


def test_comparison_renders_overlay_and_difference_highlight(tmp_path):
    def make_dxf(path: Path, shift: float) -> Path:
        doc = ezdxf.new("R2010")
        doc.header["$INSUNITS"] = 6
        polyline = doc.modelspace().add_lwpolyline(
            [(shift, 0), (shift + 10, 0), (shift + 10, 10), (shift, 10)],
            dxfattribs={"layer": "CONCEPT_BUILDING"},
        )
        polyline.close(True)
        doc.saveas(path)
        return path

    dxf_a = make_dxf(tmp_path / "scheme_a.dxf", 0)
    dxf_b = make_dxf(tmp_path / "scheme_b.dxf", 2)
    project_a = _write_project(
        tmp_path / "scheme_a.ptx",
        "方案 A",
        {"task_type": "concept_plan", "output_files": [["DXF", str(dxf_a)]]},
    )
    project_b = _write_project(
        tmp_path / "scheme_b.ptx",
        "方案 B",
        {"task_type": "concept_plan", "output_files": [["DXF", str(dxf_b)]]},
    )

    figure = Figure(figsize=(6, 4))
    records, errors = render_project_overlays(figure, [project_a, project_b])

    assert not errors
    assert len(records) == 2
    assert records[0]["unique_area"] > 0
    assert records[1]["unique_area"] > 0
    assert len(figure.axes) == 1
    assert len(figure.axes[0].patches) >= 4

    output = export_overlay_png(tmp_path / "exports", figure)
    assert output.is_file()
    assert output.stat().st_size > 0
