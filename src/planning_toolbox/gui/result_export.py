"""Export the current analysis result into student-friendly deliverables."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
import re
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.lib.pagesizes import A4


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S_%f")


def _safe_name(value: str) -> str:
    cleaned = re.sub(r"[<>:\\/?*|\"\x00-\x1f]+", "_", value).strip(" .")
    return cleaned or "分析结果"


def _scalar(value: Any) -> str:
    if isinstance(value, (list, tuple, set)):
        return "；".join(str(item) for item in value)
    if isinstance(value, dict):
        return "; ".join(f"{key}={val}" for key, val in value.items())
    return str(value)


def _detail_rows(result: Dict[str, Any]) -> List[Tuple[str, List[Dict[str, Any]]]]:
    rows: List[Tuple[str, List[Dict[str, Any]]]] = []
    for key, title in (
        ("indicators", "指标明细"),
        ("setback_results", "退线检查明细"),
        ("items", "批量处理明细"),
    ):
        values = result.get(key)
        if not isinstance(values, list) or not values or not all(isinstance(item, dict) for item in values):
            continue
        rows.append((title, values))
    return rows


def _base_name(output_dir: Path, task_name: str, suffix: str) -> Path:
    return output_dir / f"{_safe_name(task_name)}_{_timestamp()}.{suffix}"


def export_result_excel(output_dir: Path | str, task_name: str, result: Dict[str, Any]) -> Path:
    """Write a readable XLSX workbook containing summaries and detail tables."""
    output_path = Path(output_dir).resolve()
    output_path.mkdir(parents=True, exist_ok=True)
    path = _base_name(output_path, task_name, "xlsx")

    workbook = Workbook()
    summary = workbook.active
    summary.title = "结果摘要"
    header_fill = PatternFill("solid", fgColor="DCE6F0")
    header_font = Font(bold=True, color="3E536E")
    summary.append(["Planning Toolbox 分析结果", ""])
    summary["A1"].font = Font(bold=True, size=14, color="566D8E")
    summary.append(["任务", task_name])
    for key, value in result.items():
        if key in {"output_files", "indicators", "setback_results", "items"}:
            continue
        summary.append([key, _scalar(value)])
    summary.append([])
    summary.append(["生成文件", "路径"])
    for cell in summary[summary.max_row]:
        cell.fill = header_fill
        cell.font = header_font
    for label, file_path in result.get("output_files", []):
        summary.append([label, str(file_path)])
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 80
    summary.freeze_panes = "A3"

    for title, records in _detail_rows(result):
        sheet = workbook.create_sheet(title[:31])
        keys: List[str] = []
        for record in records:
            for key in record:
                if key not in keys:
                    keys.append(key)
        sheet.append(keys)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for record in records:
            sheet.append([_scalar(record.get(key, "")) for key in keys])
        for column in sheet.columns:
            column_letter = column[0].column_letter
            sheet.column_dimensions[column_letter].width = min(
                42, max(12, max(len(str(cell.value or "")) for cell in column) + 2)
            )
        sheet.freeze_panes = "A2"

    workbook.save(path)
    return path


def _pdf_lines(task_name: str, result: Dict[str, Any]) -> Iterable[str]:
    yield "Planning Toolbox 分析结果报告"
    yield f"任务：{task_name}"
    yield ""
    for key, value in result.items():
        if key in {"output_files", "indicators", "setback_results", "items"}:
            continue
        yield f"{key}：{_scalar(value)}"
    yield ""
    yield "生成文件："
    for label, file_path in result.get("output_files", []):
        yield f"- {label}：{file_path}"
    for title, records in _detail_rows(result):
        yield ""
        yield f"{title}：共 {len(records)} 条"
        for index, record in enumerate(records[:12], start=1):
            compact = "；".join(f"{key}={_scalar(value)}" for key, value in record.items())
            yield f"{index}. {compact}"
        if len(records) > 12:
            yield "（明细过多，完整内容请查看 Excel 文件。）"
    yield ""
    yield "使用边界：本报告用于学习、方案比较和结果整理，不能替代当地规划条件、正式规范核对或人工专业复核。"


def export_result_pdf(output_dir: Path | str, task_name: str, result: Dict[str, Any]) -> Path:
    """Write a Chinese-readable PDF summary using a built-in CJK font."""
    output_path = Path(output_dir).resolve()
    output_path.mkdir(parents=True, exist_ok=True)
    path = _base_name(output_path, task_name, "pdf")

    font_name = "STSong-Light"
    if font_name not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    document = pdf_canvas.Canvas(str(path), pagesize=A4)
    width, height = A4
    document.setTitle(f"Planning Toolbox - {task_name}")
    y = height - 48
    document.setFont(font_name, 10)
    for raw_line in _pdf_lines(task_name, result):
        line = raw_line or " "
        while len(line) > 54:
            document.drawString(38, y, line[:54])
            line = "  " + line[54:]
            y -= 16
            if y < 48:
                document.showPage()
                document.setFont(font_name, 10)
                y = height - 48
        document.drawString(38, y, line)
        y -= 16
        if y < 48:
            document.showPage()
            document.setFont(font_name, 10)
            y = height - 48
    document.save()
    return path


def export_preview_png(output_dir: Path | str, task_name: str, preview_canvas: Any) -> Path:
    """Save the current 2D CAD preview as a presentation-friendly PNG."""
    output_path = Path(output_dir).resolve()
    output_path.mkdir(parents=True, exist_ok=True)
    path = _base_name(output_path, task_name, "png")
    save_png = getattr(preview_canvas, "save_png", None)
    if callable(save_png):
        return Path(save_png(path, max_dimension=2400))
    preview_canvas.figure.savefig(
        path,
        dpi=180,
        bbox_inches="tight",
        facecolor=preview_canvas.figure.get_facecolor(),
    )
    return path


def export_result_artifacts(
    output_dir: Path | str,
    task_name: str,
    result: Dict[str, Any],
    preview_canvas: Any,
) -> List[Tuple[str, Path]]:
    """Create Excel, PDF and PNG deliverables for the current result."""
    return [
        ("Excel 结果表", export_result_excel(output_dir, task_name, result)),
        ("PDF 结果报告", export_result_pdf(output_dir, task_name, result)),
        ("PNG 预览图", export_preview_png(output_dir, task_name, preview_canvas)),
    ]
