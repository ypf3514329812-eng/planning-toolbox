"""Extract and export comparable indicators from saved coursework projects."""

from __future__ import annotations

from datetime import datetime
import csv
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from planning_toolbox.gui.project_file import load_project


COMPARISON_COLUMNS = [
    "方案名称",
    "任务类型",
    "输入图纸",
    "有效地块/处理数",
    "总用地面积(m²)",
    "FAR",
    "建筑密度(%)",
    "绿地率(%)",
    "停车位",
    "退线/标准化结果",
    "项目文件",
]


def _number(value: Any, default: float | None = None) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return default
    return number


def _mean(values: Iterable[Any]) -> float | None:
    numbers = [number for number in (_number(value) for value in values) if number is not None]
    return sum(numbers) / len(numbers) if numbers else None


def _format_range(values: Iterable[Any]) -> str:
    numbers = [number for number in (_number(value) for value in values) if number is not None]
    if not numbers:
        return "—"
    low, high = min(numbers), max(numbers)
    return f"{low:.2f}" if abs(low - high) < 1e-9 else f"{low:.2f} ~ {high:.2f}"


def _task_label(task_type: str) -> str:
    return {
        "parcel": "地块面积与编号",
        "indicator": "规划指标计算",
        "validate": "拓扑与退线检查",
        "gis_export": "CAD 导出 GIS",
        "gis_import": "GIS 导入 CAD",
        "batch": "批量分析",
        "concept_plan": "概念方案草图",
        "layer_standardize": "CAD 图层标准化",
    }.get(task_type, task_type or "未命名任务")


def build_comparison_row(project_path: Path | str) -> Dict[str, Any]:
    """Read one .ptx project and derive a transparent comparison row."""
    path = Path(project_path).resolve()
    state = load_project(path)
    result = state.get("last_result") or {}
    task_type = str(result.get("task_type", ""))
    indicators = result.get("indicators") or []
    row: Dict[str, Any] = {
        "方案名称": path.stem,
        "任务类型": state.get("last_task_name") or _task_label(task_type),
        "输入图纸": Path(str(state.get("dxf_path", result.get("source_file", "")))).name or "—",
        "有效地块/处理数": "—",
        "总用地面积(m²)": None,
        "FAR": "—",
        "建筑密度(%)": None,
        "绿地率(%)": None,
        "停车位": "—",
        "退线/标准化结果": "—",
        "项目文件": str(path),
    }

    if task_type == "parcel":
        row["有效地块/处理数"] = result.get("valid_count", "—")
        row["总用地面积(m²)"] = _number(result.get("total_m2"))
    elif task_type == "indicator":
        row["有效地块/处理数"] = len(indicators)
        row["总用地面积(m²)"] = sum(
            number for number in (_number(item.get("site_area_m2")) for item in indicators) if number is not None
        ) or None
        row["FAR"] = _format_range(item.get("far") for item in indicators)
        row["建筑密度(%)"] = _mean(item.get("building_density_pct") for item in indicators)
        row["绿地率(%)"] = _mean(item.get("green_ratio_pct") for item in indicators)
    elif task_type == "concept_plan":
        row["有效地块/处理数"] = result.get("parcels_count", "—")
        row["总用地面积(m²)"] = _number(result.get("parcel_area_m2"))
        estimated_far = _number(result.get("estimated_far"))
        row["FAR"] = f"{estimated_far:.2f}" if estimated_far is not None else "—"
        coverage = _number(result.get("actual_coverage_ratio"))
        row["建筑密度(%)"] = coverage * 100 if coverage is not None else None
        area = _number(result.get("parcel_area_m2"))
        green = _number(result.get("green_area_m2"))
        row["绿地率(%)"] = green / area * 100 if area and green is not None else None
        required = result.get("parking_required")
        generated = result.get("parking_generated")
        if required is not None and generated is not None:
            row["停车位"] = f"{generated}/{required}"
        minimum_setback = _number(result.get("minimum_setback_m"))
        row["退线/标准化结果"] = f"最小退线 {minimum_setback:.2f}m" if minimum_setback is not None else "未生成建筑"
    elif task_type == "validate":
        row["有效地块/处理数"] = result.get("valid_count", "—")
        checks = result.get("setback_results") or []
        compliant = sum(item.get("status") == "COMPLIANT" for item in checks)
        row["退线/标准化结果"] = f"合规 {compliant}/{len(checks)}" if checks else "无退线结果"
    elif task_type == "batch":
        row["有效地块/处理数"] = f"{result.get('success_count', 0)}/{result.get('processed_count', 0)}"
        total_ha = sum(_number(item.get("total_ha"), 0.0) or 0.0 for item in result.get("items", []))
        row["总用地面积(m²)"] = total_ha * 10000
    elif task_type == "layer_standardize":
        row["有效地块/处理数"] = result.get("remapped_total", "—")
        unmapped = len(result.get("unmapped_layers") or [])
        row["退线/标准化结果"] = f"已映射；未识别图层 {unmapped} 个"

    return row


def build_comparison_rows(project_paths: Iterable[Path | str]) -> Tuple[List[Dict[str, Any]], List[str]]:
    rows: List[Dict[str, Any]] = []
    errors: List[str] = []
    for project_path in project_paths:
        try:
            rows.append(build_comparison_row(project_path))
        except Exception as exc:
            errors.append(f"{Path(project_path).name}: {exc}")
    return rows, errors


def _export_name(output_dir: Path | str, suffix: str) -> Path:
    root = Path(output_dir).resolve()
    root.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    return root / f"方案对比_{timestamp}.{suffix}"


def export_comparison_csv(output_dir: Path | str, rows: List[Dict[str, Any]]) -> Path:
    path = _export_name(output_dir, "csv")
    with path.open("w", newline="", encoding="utf-8-sig") as stream:
        writer = csv.DictWriter(stream, fieldnames=COMPARISON_COLUMNS)
        writer.writeheader()
        writer.writerows({column: row.get(column, "") for column in COMPARISON_COLUMNS} for row in rows)
    return path


def export_comparison_excel(output_dir: Path | str, rows: List[Dict[str, Any]]) -> Path:
    path = _export_name(output_dir, "xlsx")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "方案对比"
    sheet.append(COMPARISON_COLUMNS)
    fill = PatternFill("solid", fgColor="DCE6F0")
    font = Font(bold=True, color="3E536E")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
    for row in rows:
        sheet.append([row.get(column, "") if row.get(column) is not None else "—" for column in COMPARISON_COLUMNS])
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(52, max(14, max(len(str(cell.value or "")) for cell in column) + 2))
    sheet.freeze_panes = "A2"
    workbook.save(path)
    return path
